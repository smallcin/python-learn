#coding=utf-8

from openpyxl import Workbook
import json

with open("/Users/zh_wj/Personal/python-learn/test/LeagueExploreGridInfoTab.txt","rb") as loadFile:

	load_dic = json.load(loadFile)
	print "类型：{}".format(type(load_dic))
	load_dic_row = load_dic[1][3]
	print "值：{}".format(load_dic_row)
	print "sixGridStr_：{}".format(load_dic_row['sixGridStr_'])
	print "circle_:{}".format(load_dic_row['circle_'])
	print "index_:{}".format(load_dic_row['index_'])
	print "row_:{}".format(load_dic_row['row_'])

	#创建一个excel文件，写入不同类的内容
	w = Workbook() #创建文件对象
	ws = w.active #获取第一个sheet
	ws.title = "league_exploreMap"
	ws['A1'] = "公会探索地图点位" #A列第一行写入42
	ws['A2'] = "CS"
	ws['A3'] = "number"
	ws['A4'] = "number"
	ws['A5'] = "id"
	ws['B1'] = "相邻点位"
	ws['B2'] = "CS"
	ws['B3'] = "numArray"
	ws['B4'] = "string"
	ws['B5'] = "adjacentIds"
	ws['C1'] = "点位坐标"
	ws['C2'] = "C"
	ws['C3'] = "numArray"
	ws['C4'] = "string"
	ws['C5'] = "gridPos"
	ws['D1'] = "阻碍标记"
	ws['D2'] = "C"
	ws['D3'] = "number"
	ws['D4'] = "number"
	ws['D5'] = "blockType"
	ws['E1'] = "事件ID"
	ws['E2'] = "S"
	ws['E3'] = "number"
	ws['E4'] = "number"
	ws['E5'] = "eventId"

	#ws.append([1, 2, 3]) #追加行，写入多个单元格
	print ("list 长度:",len(load_dic[3]))
	index = 1
	for i in range(0,4):
		for load_dic_row_tmp in load_dic[i]:

			strTmp = "A{}".format(index+5)
			ws[strTmp] = load_dic_row_tmp['index_']
			strTmpB = "B{}".format(index+5)
			ws[strTmpB] = load_dic_row_tmp['sixGridStr_']
			strTmpC = "C{}".format(index+5)
			ws[strTmpC] = "{},{}".format(load_dic_row_tmp['circle_']['x'], load_dic_row_tmp['circle_']['y'])
			strTmpD = "D{}".format(index+5)
			ws[strTmpD] = load_dic_row_tmp['blockId_']
			strTmpE = "E{}".format(index+5)
			ws[strTmpE] = load_dic_row_tmp['eventId_']
			index = index + 1
	

	# for i in xrange(1,10):
	# 	strTmp = "A{}".format(i)
	# 	ws[strTmp] = "ID{}".format(i)

	#保存文件
	w.save('/Users/zh_wj/Personal/python-learn/test/leagueExploreAuto.xlsx')

